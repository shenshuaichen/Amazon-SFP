import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
#import codecs
import datetime
import easygui as gui
address = os.getcwd()
#import win32com.client as win32

###########################################convert .xls to .xlsx ###########################################
with open('Active Listings Report.txt','r',encoding='gbk',errors='ignore') as f:
    data=f.readlines()
print(len(data))

# create a workbook
wb3 = openpyxl.Workbook()

#open the first one workbook
ws3 = wb3.active
for i1 in data:
    ws3.append(i1.split('\t'))
wb3.save("Active Listings Report.xlsx")
#print('save')
#print('File .txt convert .xlsx successful!!')

###########################################Open Bundles.xlsx and searchresults.xlsx ##################################
df1 = pd.read_excel('Bundles.xlsx') 
df2 = pd.read_excel('searchresults.xlsx') 


df2 = df2.drop(df2.columns[[2,3]], axis = 1) #将第二个文件的3，4竖列删掉 delete column 2, 3 of searchresults.xlsx 

horizontal_stack = pd.concat([df1, df2], axis = 1) #combine Bundles.xlsx and searchresults.xlsx
horizontal_stack.insert(8,'0', 0) #add a empty column at 7th
horizontal_stack.insert(9, '1', 0) #add a empty column at 8th
horizontal_stack.insert(10, '2', 0) #add a empty column at 9th
horizontal_stack.insert(11, '3', 0) #add a empty column at 10th
horizontal_stack.insert(12, '4', 0) #add a empty column at 11th
horizontal_stack.insert(13, '5', 0) #add a empty column at 12th
horizontal_stack.insert(14, '6', 0) #add a empty column at 13th
horizontal_stack.insert(15, '7', 0) #add a empty column at 14th
horizontal_stack.insert(17, '8', 0) #add a empty column at 15th
#horizontal_stack.insert(21, '9', 0) #add a empty column at 16th

#save the file and create name test_Inventory.xlsx
export_excel = horizontal_stack.to_excel(address + '\\test_Inventory.xlsx', index = None, header = True)

###########################################Open test_Inventory.xlsx，not SFP.xlsx，active listing.xlsx ##########################    
wb = load_workbook(address + '\\test_Inventory.xlsx') 
ws = wb.active 
wb2 = load_workbook(address + '\\Not SFP.xlsx') 
ws2 = wb2.active 
wb4 = load_workbook(address + '\\Active Listings Report.xlsx')
ws4 = wb4.active
wb6 = load_workbook(address + '\\Amazon Shipping Template Empty.xlsx')
ws6 = wb6.active

###########################################calculating how many row in test_Inventory file ##########################################
num = 1
while 1: 
    cell = ws.cell(row=num, column=1).value
    if cell:
        num = num +1
    else:
        break
###########################################calculating SFP ###########################################
dic_O = {}
dic_M = {}
dic_K = {}
for q in range(ws.max_row):
    dic_O[ws["Q%d" % (q+1)].value] = ws["T%d" % (q+1)].value
    dic_M[ws["Q%d" % (q+1)].value] = ws["U%d" % (q+1)].value
    dic_K[ws["Q%d" % (q+1)].value] = ws["V%d" % (q+1)].value

i = 1
while i < num:
    if str(ws["H%d" % (i)].value).find(":") != -1:
        ws["I%d" % (i)].value = str(ws["H%d" % (i)].value).split(" : ")[1]
        ws["J%d" % (i)].value = dic_O.get(ws["I%d" % (i)].value)
        ws["K%d" % (i)].value = dic_M.get(ws["I%d" % (i)].value)
        ws["L%d" % (i)].value = dic_K.get(ws["I%d" % (i)].value)
        i = i +1
    else:
        ws["I%d" % (i)].value = ws["H%d" % (i)].value
        ws["J%d" % (i)].value = dic_O.get(ws["I%d" % (i)].value)
        ws["K%d" % (i)].value = dic_M.get(ws["I%d" % (i)].value)
        ws["L%d" % (i)].value = dic_K.get(ws["I%d" % (i)].value)
        i = i +1

dic_O_J = {}
dic_M_K = {}
dic_K_L = {}

for each in range(num):
    dic_O_J[ws["I%d" % (each + 1)].value] = ws["J%d" % (each + 1)].value
    dic_M_K[ws["I%d" % (each + 1)].value] = ws["K%d" % (each + 1)].value
    dic_K_L[ws["I%d" % (each + 1)].value] = ws["L%d" % (each + 1)].value

ws["J1"].value = 1
ws["K1"].value = 1
ws["K1"].value = 1
#################################Calculating Bundles##################################
next_number = 2
for each1 in range(num):
    if ws["A%d" % (each1 + 1)].value == None:
        break
    if ws["A%d" % (each1 + 1)].value.find("_") != -1:
        ws["M%d" % (each1 + 1)].value = ws["J%d" % (each1 + 1)].value
        ws["N%d" % (each1 + 1)].value = ws["K%d" % (each1 + 1)].value
        ws["O%d" % (each1 + 1)].value = ws["L%d" % (each1 + 1)].value
    if ws["A%d" % (each1 + 1)].value.find("+") != -1:
        list_bundles = ws["A%d" % (each1 + 1)].value.split("+")
        mini_O = []
        mini_M = []
        mini_K = []
        for each in range(len(list_bundles)):
            mini_O.append(dic_O_J.get(list_bundles[each]))
            mini_M.append(dic_M_K.get(list_bundles[each]))
            mini_K.append(dic_K_L.get(list_bundles[each]))
            ws["M%d" % (each1 + 1)].value = min(mini_O)
            ws["N%d" % (each1 + 1)].value = min(mini_M)
            ws["O%d" % (each1 + 1)].value = min(mini_K)
    else:
        ontaio_a = []
        memphis_a = []
        Kansas_a = []
        ontaio_a.append(ws["J%d" % (each1 + 1)].value)
        memphis_a.append(ws["K%d" % (each1 + 1)].value)
        Kansas_a.append(ws["L%d" % (each1 + 1)].value)
        if ws["A%d" % (each1 + next_number)].value == ws["A%d" % (each1 + 1)].value:
            ontaio_a.append(ws["J%d" % (each1 + next_number)].value)
            memphis_a.append(ws["K%d" % (each1 + next_number)].value)
            Kansas_a.append(ws["L%d" % (each1 + next_number)].value)
            ws["M%d" % (each1 + 1)].value = min(ontaio_a)
            ws["N%d" % (each1 + 1)].value = min(memphis_a)
            ws["O%d" % (each1 + 1)].value = min(Kansas_a)
            ws["M%d" % (each1 + next_number)].value = min(ontaio_a)
            ws["N%d" % (each1 + next_number)].value = min(memphis_a)
            ws["O%d" % (each1 + next_number)].value = min(Kansas_a)


ws["M1"] = 0
ws["N1"] = 0
ws["O1"] = 0
ws["P1"] = 0

for p in range(num):
    if ws['M%d' % (p + 1)].value >= 5 and ws['N%d' % (p + 1)].value >= 5 and ws['O%d' % (p + 1)].value >= 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Complete'
    if ws['M%d' % (p + 1)].value >= 5 and ws['N%d' % (p + 1)].value >= 5 and ws['O%d' % (p + 1)].value < 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Ontario + Memphis'
    if ws['M%d' % (p + 1)].value >= 5 and ws['N%d' % (p + 1)].value < 5 and ws['O%d' % (p + 1)].value >= 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Ontario + Kansas City'
    if ws['M%d' % (p + 1)].value < 5 and ws['N%d' % (p + 1)].value >= 5 and ws['O%d' % (p + 1)].value >= 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Kansas City + Memphis'
    if ws['M%d' % (p + 1)].value < 5 and ws['N%d' % (p + 1)].value < 5 and ws['O%d' % (p + 1)].value < 5:
        ws['P%d' % (p + 1)].value = 'Default Amazon Template'
    if ws['M%d' % (p + 1)].value >= 5 and ws['N%d' % (p + 1)].value < 5 and ws['O%d' % (p + 1)].value < 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Ontario'
    if ws['M%d' % (p + 1)].value < 5 and ws['N%d' % (p + 1)].value >= 5 and ws['O%d' % (p + 1)].value < 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Memphis'
    if ws['M%d' % (p + 1)].value < 5 and ws['N%d' % (p + 1)].value < 5 and ws['O%d' % (p + 1)].value >= 5:
        ws['P%d' % (p + 1)].value = 'Seller Fulfilled Prime - Kansas City'
    if ws['M%d' % (p + 1)].value is None or ws['N%d' % (p + 1)].value is None or ws['O%d' % (p + 1)].value is None:
        break


ws['P1'].value = 'SFP'
ws["H1"].value = "Name.1"
ws["I1"].value = "单品编号"
ws["J1"].value = "单品库存：Ontraio"
ws["K1"].value = "单品库存：Memphis"
ws["L1"].value = "单品库存：Kansas City"
ws["M1"].value = "组合库存：Ontario"
ws["N1"].value = "组合库存：Memphis"
ws["O1"].value = "组合库存：Kansas City"

###########################################Calculating single SFP ###########################################
ws['T1'].value = 0
ws['U1'].value = 0
ws['V1'].value = 0
for v in range(ws.max_row):
    try:
        if ws['T%d' % (v + 1)].value >= 5 and ws['U%d' % (v + 1)].value >= 5 and ws['V%d' % (v + 1)].value >= 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Complete'
        if ws['T%d' % (v + 1)].value >= 5 and ws['U%d' % (v + 1)].value >= 5 and ws['V%d' % (v + 1)].value < 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Ontario + Memphis'
        if ws['T%d' % (v + 1)].value >= 5 and ws['U%d' % (v + 1)].value < 5 and ws['V%d' % (v + 1)].value >= 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Ontario + Kansas City'
        if ws['T%d' % (v + 1)].value < 5 and ws['U%d' % (v + 1)].value >= 5 and ws['V%d' % (v + 1)].value >= 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Kansas City + Memphis'
        if ws['T%d' % (v + 1)].value < 5 and ws['U%d' % (v + 1)].value < 5 and ws['V%d' % (v + 1)].value < 5:
            ws['W%d' % (v + 1)].value = 'Default Amazon Template'
        if ws['T%d' % (v + 1)].value >= 5 and ws['U%d' % (v + 1)].value < 5 and ws['V%d' % (v + 1)].value < 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Ontario'
        if ws['T%d' % (v + 1)].value < 5 and ws['U%d' % (v + 1)].value >= 5 and ws['V%d' % (v + 1)].value < 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Memphis'
        if ws['T%d' % (v + 1)].value < 5 and ws['U%d' % (v + 1)].value < 5 and ws['V%d' % (v + 1)].value >= 5:
            ws['W%d' % (v + 1)].value = 'Seller Fulfilled Prime - Kansas City' 
    except:
        break
ws['T1'].value = 'Ontario'
ws['U1'].value = 'Memphis'
ws['V1'].value = 'Kansas City'
#print('Single products SFP calculates successful！')

###########################################calculating some products cannot SFP ###########################################
rowlist = []
for row in ws2.rows:
    for cell in row:
        cellValue = cell.value
        rowlist.append(cellValue)
for w in range(ws.max_row):
    if ws['Q%d' % (w + 1)].value in rowlist:
        ws['W%d' % (w + 1)].value = 'Default Amazon Template'
ws['W1'].value = '单品SFP'
#print('Not SFP calculates successful！')

###########################################Calculating quantity ###########################################
active_list_report = {}
for d in range(ws4.max_row):
    active_list_report[ws4["D%d" % (d + 1)].value] = ws4["F%d" % (d + 1)].value
for c in range(ws.max_row):
    ws["C%d" % (c + 1)].value = active_list_report.get(ws["A%d" % (c + 1)].value)
    ws["S%d" % (c + 1)].value = active_list_report.get(ws["Q%d" % (c + 1)].value)
ws["C1"].value = "Quantity"
ws["S1"].value = "quantity"

###########################################Add item-name###########################################
dic_item_name = {}
for b in range(ws4.max_row):
    dic_item_name[ws4["D%d" % (b +1)].value] = ws4["A%d" % (b + 1)].value
for b1 in range(num):
    ws["B%d" % (b1 + 1)].value = dic_item_name.get(ws["Q%d" % (b1 + 1)].value)
for r in range(ws.max_row):
    ws["R%d" % (r + 1)].value = dic_item_name.get(ws["Q%d" % (r + 1)].value)


ws["B1"].value = "Item-name"
ws["R1"].value =  "item-name"

###########################################Import Amazon Shipping Template #####################################################
Name_sku = []
item_name_list = []
SFP_list = []

for amazonlist in range(num):
    Name_sku.append(ws["A%d" % (amazonlist + 1)].value)
    item_name_list.append(ws["B%d" % (amazonlist + 1)].value)
    SFP_list.append(ws["P%d" % (amazonlist + 1)].value)


for amazonlist2 in range(ws.max_row):
    Name_sku.append(ws["Q%d" % (amazonlist2 + 1)].value)
    item_name_list.append(ws["R%d" % (amazonlist2 + 1)].value)
    SFP_list.append(ws["W%d" % (amazonlist2 + 1)].value)

Name_sku.remove("Name")
Name_sku.remove("SKU")
item_name_list.remove("Item-name")
item_name_list.remove("item-name")
SFP_list.remove("SFP")
SFP_list.remove("单品SFP")

x = 4
for amazonlist3 in Name_sku:
    ws6["B%d" % (x)].value = amazonlist3
    x +=1
y = 4
for amazonlist4 in item_name_list:
    ws6["F%d" % (y)].value = amazonlist4
    y += 1
z = 4
for amazonlist5 in SFP_list:
    ws6["L%d" % (z)].value = amazonlist5
    z += 1


###########################################Save#####################################################
today_time = datetime.date.today()
wb.save(address + '\\test_Inventory.xlsx')
wb6.save(address + '\\Amazon Shipping Template-%s.xlsx' % (today_time))
gui.msgbox('Done!!')
